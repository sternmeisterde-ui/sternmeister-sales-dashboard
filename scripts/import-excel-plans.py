#!/usr/bin/env python3
"""
Import historical Daily plans & facts from Excel sheets into the daily_plans
table.

Source files (repo root):
  - "дейли коммерция.xlsx"             — B2B
  - "Госники Daily Weekly Monthly (1).xlsx" — B2G

Scope: Jan 1 2026 → today. Writes to `daily_plans` at periodType='day'
(per-day values from Daily_Numbers) and periodType='month' (from Monthly
Numbers). Week rollups are computed from daily values for ISO weeks that
overlap the target range. userId=NULL (line-level default plans).

Idempotent: UPSERT on (department, line, metric_key, period_type, period_date,
user_id). Re-running overwrites planValue for the same key but never deletes.

Read-only side-effects: no destructive DDL. Safe to run multiple times.
"""
from __future__ import annotations

import os
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2
from psycopg2.extras import execute_values


ROOT = Path(__file__).resolve().parent.parent
B2B_FILE = ROOT / "дейли коммерция.xlsx"
B2G_FILE = ROOT / "Госники Daily Weekly Monthly (1).xlsx"

START_DATE = date(2026, 1, 1)
END_DATE = date.today()


def load_env() -> str:
    env_path = ROOT / ".env.local"
    if not env_path.exists():
        env_path = ROOT / ".env"
    url = None
    for ln in env_path.read_text().splitlines():
        if ln.startswith("DATABASE_URL="):
            url = ln.split("=", 1)[1].strip().strip('"').strip("'")
            break
    if not url:
        sys.exit("DATABASE_URL not found in .env.local / .env")
    return url


# ─────────────────────────────────────────────────────────────────────
# B2B mapping — Daily_Numbers (day-per-column) + Monthly Numbers
# ─────────────────────────────────────────────────────────────────────

# (metricKey, line, excel_daily_row, excel_monthly_row, pct_scale)
#   pct_scale: True → Excel stores 0..1 but we want 0..100 (percentages)
B2B_METRICS: list[tuple[str, str, int | None, int | None, bool]] = [
    # salesTotal section
    ("total_revenueTotal_p",    "salesTotal", 47,  20, False),
    ("total_revenueTotal_f",    "salesTotal", 48,  21, False),
    ("total_newRevenue_p",      "salesTotal", 49,  20, False),   # same as revenueTotal at team level
    ("total_newRevenue_f",      "salesTotal", 50,  21, False),
    ("total_komLeads_p",        "salesTotal", 51,  None, False),
    ("total_komLeads_f",        "salesTotal", 52,  None, False),
    ("total_sales_p",           "salesTotal", 53,  None, False),
    ("total_sales_f",           "salesTotal", 54,  None, False),
    ("total_prepayments",       "salesTotal", 55,  None, False),
    ("total_ql2p_p",            "salesTotal", 56,  None, True),
    ("total_ql2p_f",            "salesTotal", 57,  None, True),
    ("total_avgCheck_p",        "salesTotal", 58,  None, False),
    ("total_avgCheck_f",        "salesTotal", 59,  None, False),
    # Бух section
    ("buh_salesPlusRenewals_p", "salesBuh",   63,  22, False),
    ("buh_salesPlusRenewals_f", "salesBuh",   64,  23, False),
    ("buh_newRevenue_p",        "salesBuh",   65,  25, False),
    ("buh_newRevenue_f",        "salesBuh",   66,  26, False),
    ("buh_komLeads_p",          "salesBuh",   67,  27, False),
    ("buh_komLeads_f",          "salesBuh",   68,  28, False),
    ("buh_sales_p",             "salesBuh",   71,  31, False),
    ("buh_sales_f",             "salesBuh",   72,  32, False),
    ("buh_prepayments",         "salesBuh",   73,  33, False),
    ("buh_ql2p_p",              "salesBuh",   74,  35, True),
    ("buh_ql2p_f",              "salesBuh",   75,  36, True),
    ("buh_avgCheck_p",          "salesBuh",   78,  39, False),
    ("buh_avgCheck_f",          "salesBuh",   79,  40, False),
    # Мед section
    ("med_salesPlusRenewals_p", "salesMed",   83,  46, False),
    ("med_salesPlusRenewals_f", "salesMed",   84,  47, False),
    ("med_newRevenue_p",        "salesMed",   85,  48, False),
    ("med_newRevenue_f",        "salesMed",   86,  49, False),
    ("med_komLeads_p",          "salesMed",   87,  50, False),
    ("med_komLeads_f",          "salesMed",   88,  51, False),
    ("med_sales_p",             "salesMed",   89,  52, False),
    ("med_sales_f",             "salesMed",   90,  53, False),
    ("med_prepayments",         "salesMed",   91,  54, False),
    ("med_ql2p_p",              "salesMed",   92,  55, True),
    ("med_ql2p_f",              "salesMed",   93,  56, True),
    ("med_avgCheck_p",          "salesMed",   94,  57, False),
    ("med_avgCheck_f",          "salesMed",   95,  58, False),
    # Calls + OKK section
    ("calls_managersOnLine_p",  "calls",      119, 65, False),
    ("calls_managersOnLine_f",  "calls",      120, 66, False),
    ("calls_total_p",           "calls",      121, 67, False),
    ("calls_total_f",           "calls",      122, 68, False),
    ("calls_totalMinutes_p",    "calls",      123, 69, False),
    ("calls_totalMinutes_f",    "calls",      124, 70, False),
    ("calls_avgWait_p",         "calls",      125, 71, False),
    ("calls_avgWait_f",         "calls",      126, 72, False),
    ("calls_dialPercent_p",     "calls",      127, 73, True),
    ("calls_dialPercent_f",     "calls",      128, 74, True),
    ("calls_sla_p",             "calls",      129, 75, False),
    ("calls_sla_f",             "calls",      130, 76, False),
    ("okk_buh1_p",              "calls",      131, 77, True),
    ("okk_buh1_f",              "calls",      132, 78, True),
    ("okk_buh2_p",              "calls",      133, 79, True),
    ("okk_buh2_f",              "calls",      134, 80, True),
    ("okk_med1_p",              "calls",      135, None, True),
    ("okk_med1_f",              "calls",      136, None, True),
    ("okk_avg_p",               "calls",      137, 81, True),
    ("okk_avg_f",               "calls",      138, 82, True),
]

# B2B Daily_Numbers: column offsets for each month start (verified via row 1)
B2B_MONTH_COL: dict[tuple[int, int], int] = {
    (2026, 1): 277,
    (2026, 2): 308,
    (2026, 3): 336,
    (2026, 4): 367,
}

# B2B Monthly Numbers: one column per month (row 1 labels already verified)
B2B_MONTHLY_COL: dict[tuple[int, int], int] = {
    (2026, 1): 11,
    (2026, 2): 12,
    (2026, 3): 13,
    (2026, 4): 14,
}


# ─────────────────────────────────────────────────────────────────────
# B2G mapping — Daily_Numbers (3 lines in one sheet) + Monthly Numbers
# ─────────────────────────────────────────────────────────────────────

# B2G Daily_Numbers has three nested "Звонки" blocks:
#   Менеджер-квалификатор (line 1):  R36 header, metrics R37–R58
#   Доведение (line 3):              R59 header, metrics R60–R81
#   Менеджер второй линии (line 2):  R82 header, metrics R83–R104
# We map common call metrics to each line.

def b2g_calls_row(line: str, label: str) -> int | None:
    """Map a call metric label → daily row number for the given B2G line."""
    # Rows relative to each line's header (subtract header_row)
    rows_by_line = {
        "1": {   # Менеджер-квалификатор (header R36)
            "staffCount": 38, "callsTotal_p": 39, "callsTotal": 40, "callsConnected": 41,
            "totalMinutes_p": 42, "totalMinutes": 43, "avgDialogPerEmployee": 44,
            "avgDialogMinutes": 45, "avgWait_p": 46, "avgWait_f": 47, "dialPercent": 48,
            "missedIncoming": 49, "overdueTasks": 50, "regulationPercent": 51,
            "avgCallsPerLead": 52, "sla_p": 53, "sla_f": 54, "okk_p": 55, "okk_f": 56,
            "roleplay_p": 57, "roleplay_f": 58,
        },
        "3": {   # Доведение (header R59)
            "staffCount": 60, "callsTotal_p": 61, "callsTotal": 62, "callsConnected": 63,
            "totalMinutes_p": 64, "totalMinutes": 65, "avgDialogPerEmployee": 66,
            "avgDialogMinutes": 67, "avgWait_p": 68, "avgWait_f": 69, "dialPercent": 70,
            "missedIncoming": 71, "overdueTasks": 72, "regulationPercent": 73,
            "avgCallsPerLead": 74, "sla_p": 75, "sla_f": 76, "tlt_f": 77,
            "okk_p": 78, "okk_f": 79, "roleplay_p": 80, "roleplay_f": 81,
        },
        "2": {   # Менеджер второй линии (header R82)
            "staffCount": 83, "callsTotal_p": 84, "callsTotal": 85, "callsConnected": 86,
            "totalMinutes_p": 87, "totalMinutes": 88, "avgDialogPerEmployee": 89,
            "avgDialogMinutes": 90, "avgWait_p": 91, "avgWait_f": 92, "dialPercent": 93,
            "missedIncoming": 94, "overdueTasks": 95, "regulationPercent": 96,
            "avgCallsPerLead": 97, "tlt_f": 98, "sla_p": 99, "sla_f": 100,
            "okk_p": 101, "okk_f": 102, "roleplay_p": 103, "roleplay_f": 104,
        },
    }
    return rows_by_line.get(line, {}).get(label)


# B2G line-1 funnel metrics (per-line body of Daily_Numbers, R3..R35)
B2G_FUNNEL_METRICS: list[tuple[str, int | None, int | None, bool]] = [
    # (metricKey, daily_row, monthly_row, pct_scale)
    ("activeDeals",        3,  5,  False),
    ("managersOnLine",     4,  6,  False),
    ("totalLeads_p",       5,  7,  False),
    ("totalLeads",         6,  8,  False),
    ("qualLeads_p",        7,  9,  False),
    ("qualLeads",          8,  10, False),
    ("qualLeadsPercent",   9,  11, True),
    ("a2",                 10, 14, False),
    ("b1",                 11, 15, False),
    ("b2plus",             12, 16, False),
    ("avgPortfolio",       13, 17, False),
    ("tasksTotal",         14, None, False),
    ("tasksNew",           15, None, False),
    ("convQualTask",       16, None, True),
    ("consultTotal",       17, None, False),
    ("consultNew",         18, None, False),
    ("convTaskConsult",    19, None, True),
    ("termsTotal",         20, 26, False),
    ("termsNew",           21, 27, False),
    ("awaitTermTotal",     22, 28, False),
    ("awaitTermNew",       23, 29, False),
    ("convConsultTerm",    24, 35, True),
    ("termDCCancelled",    25, 37, False),
    ("termDCDone",         26, 38, False),
    ("termAATransferred",  27, 39, False),
    ("termAACancelled",    28, 40, False),
    ("termAACount",        29, 41, False),
    ("beraterReview",      30, 42, False),
    ("delayedStart",       31, 43, False),
    ("appeal",             32, 44, False),
    ("gutscheinsApproved", 33, 46, False),
    ("beraterReject",      34, 51, False),
    ("appealsSubmitted",   35, 52, False),
]

# B2G calls metrics we care about (subset — the ones in metrics-config.ts)
B2G_CALL_METRIC_KEYS = [
    "staffCount", "callsTotal_p", "callsTotal", "callsConnected", "dialPercent",
    "missedIncoming", "totalMinutes_p", "totalMinutes", "avgDialogPerEmployee",
    "avgDialogMinutes", "avgWait_p", "overdueTasks", "regulationPercent",
    "avgCallsPerLead", "sla_p", "sla_f", "sla_shift_f", "tlt_f",
    "okk_p", "okk_f", "roleplay_p", "roleplay_f",
]

# B2G Daily_Numbers col offsets (row 2 shows date, col2=Oct 1 2025)
# We compute per-day column dynamically via date math below.
B2G_DAILY_START_DATE = date(2025, 10, 1)
B2G_DAILY_START_COL = 2
B2G_DAILY_END_COL = 213

# B2G Monthly Numbers: one col per month (row 1 labels verified)
B2G_MONTHLY_COL: dict[tuple[int, int], int] = {
    (2026, 1): 6,
    (2026, 2): 7,
    (2026, 3): 8,
    (2026, 4): 9,
}


# ─────────────────────────────────────────────────────────────────────
# Value coercion + helpers
# ─────────────────────────────────────────────────────────────────────

def coerce(v: Any, pct_scale: bool) -> str | None:
    """Convert an Excel cell to a numeric string suitable for daily_plans.
    Returns None if cell is empty or can't be parsed as a number."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#"):  # #DIV/0! etc.
            return None
        # European decimals: "10,5"
        s = s.replace(" ", "").replace("€", "").replace("$", "").replace(",", ".")
        try:
            num = float(s)
        except ValueError:
            return None
    elif isinstance(v, (int, float)):
        num = float(v)
    else:
        return None
    if pct_scale:
        num *= 100
    # Round to 2 decimals; strip trailing zeros
    if abs(num - round(num)) < 1e-9:
        return str(int(round(num)))
    return f"{num:.2f}".rstrip("0").rstrip(".")


def b2b_daily_col(year: int, month: int, day: int) -> int | None:
    base = B2B_MONTH_COL.get((year, month))
    if base is None:
        return None
    return base + (day - 1)


def b2g_daily_col(d: date) -> int | None:
    delta = (d - B2G_DAILY_START_DATE).days
    if delta < 0:
        return None
    col = B2G_DAILY_START_COL + delta
    if col > B2G_DAILY_END_COL:
        return None
    return col


def date_range(start: date, end: date):
    cur = start
    while cur <= end:
        yield cur
        cur += timedelta(days=1)


def iso_week_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


# ─────────────────────────────────────────────────────────────────────
# Pure-fact keys (hasPlan:false, hasFact:true) — these come from analytics.*
# and must NOT be written to daily_plans. Keeps the table clean so our
# override logic isn't shadowing live SQL.
# ─────────────────────────────────────────────────────────────────────

B2B_PURE_FACT_SKIP = {
    "buh_avgCheck_f", "buh_komLeads_f", "buh_planDoneNew", "buh_planDoneTotal",
    "buh_prepayments", "buh_ql2p_f", "buh_sales_f", "calls_dialPercent_f",
    "calls_managersOnLine_f", "calls_sla_f", "calls_totalMinutes_f", "calls_total_f",
    "med_avgCheck_f", "med_komLeads_f", "med_planDoneNew", "med_planDoneTotal",
    "med_prepayments", "med_ql2p_f", "med_sales_f", "okk_avg_f", "okk_buh1_f",
    "okk_buh2_f", "okk_med1_f", "total_avgCheck_f", "total_komLeads_f",
    "total_planDoneNew", "total_planDoneTotal", "total_prepayments",
    "total_ql2p_f", "total_sales_f",
}
B2G_PURE_FACT_SKIP = {
    "a2", "activeDeals", "appeal", "appealsSubmitted", "avgCallsPerLead",
    "avgDialogMinutes", "avgDialogPerEmployee", "avgPortfolio", "avgWait_p",
    "awaitTermNew", "awaitTermTotal", "b1", "b2plus", "beraterReject",
    "beraterReview", "callsConnected", "callsTotal", "callsTotal_p",
    "consultNew", "consultTotal", "convConsultTerm", "convQualTask",
    "convTaskConsult", "delayedStart", "dialPercent", "gutscheinsApproved",
    "managersOnLine", "missedIncoming", "okk_f", "okk_p", "overdueTasks",
    "qualLeads", "qualLeadsPercent", "revenue", "roleplay_f", "roleplay_p",
    "sla_f", "sla_p", "sla_shift_f", "staffCount", "tasksNew", "tasksTotal",
    "termAACancelled", "termAACount", "termAATransferred", "termDCCancelled",
    "termDCDone", "termsNew", "termsTotal", "tlt_f", "totalLeads",
    "totalMinutes", "totalMinutes_p",
}


# ─────────────────────────────────────────────────────────────────────
# Extraction
# ─────────────────────────────────────────────────────────────────────

def extract_b2b_rows() -> list[tuple]:
    wb = openpyxl.load_workbook(B2B_FILE, data_only=True)
    daily = wb["Daily_Numbers"]
    monthly = wb["Monthly Numbers"]
    rows: list[tuple] = []

    # ── Daily ──
    for d in date_range(START_DATE, END_DATE):
        col = b2b_daily_col(d.year, d.month, d.day)
        if col is None:
            continue
        for metric_key, line, drow, _mrow, pct in B2B_METRICS:
            if drow is None or metric_key in B2B_PURE_FACT_SKIP:
                continue
            v = daily.cell(drow, col).value
            val = coerce(v, pct)
            if val is None:
                continue
            rows.append(("b2b", line, None, metric_key, val, "day", d.isoformat()))

    # ── Monthly ──
    for (yr, mo), col in B2B_MONTHLY_COL.items():
        if date(yr, mo, 1) < date(START_DATE.year, START_DATE.month, 1):
            continue
        period_date = f"{yr}-{mo:02d}"
        for metric_key, line, _drow, mrow, pct in B2B_METRICS:
            if mrow is None or metric_key in B2B_PURE_FACT_SKIP:
                continue
            v = monthly.cell(mrow, col).value
            val = coerce(v, pct)
            if val is None:
                continue
            rows.append(("b2b", line, None, metric_key, val, "month", period_date))

    return rows


def extract_b2g_rows() -> list[tuple]:
    wb = openpyxl.load_workbook(B2G_FILE, data_only=True)
    daily = wb["Daily_Numbers"]
    monthly = wb["Monthly Numbers"]
    rows: list[tuple] = []

    # ── Daily: funnel (line 1 only — Excel stores only one funnel block) ──
    for d in date_range(START_DATE, END_DATE):
        col = b2g_daily_col(d)
        if col is None:
            continue
        for metric_key, drow, _mrow, pct in B2G_FUNNEL_METRICS:
            if drow is None or metric_key in B2G_PURE_FACT_SKIP:
                continue
            v = daily.cell(drow, col).value
            val = coerce(v, pct)
            if val is None:
                continue
            rows.append(("b2g", "1", None, metric_key, val, "day", d.isoformat()))

        # Calls — each of 3 lines has its own block
        for line in ("1", "2", "3"):
            for key in B2G_CALL_METRIC_KEYS:
                if key in B2G_PURE_FACT_SKIP:
                    continue
                drow = b2g_calls_row(line, key)
                if drow is None:
                    continue
                v = daily.cell(drow, col).value
                pct = key.endswith("Percent") or key in {"dialPercent", "okk_p", "okk_f", "roleplay_p", "roleplay_f", "regulationPercent"}
                val = coerce(v, pct)
                if val is None:
                    continue
                rows.append(("b2g", line, None, key, val, "day", d.isoformat()))

    # ── Monthly ──
    for (yr, mo), col in B2G_MONTHLY_COL.items():
        if date(yr, mo, 1) < date(START_DATE.year, START_DATE.month, 1):
            continue
        period_date = f"{yr}-{mo:02d}"
        for metric_key, _drow, mrow, pct in B2G_FUNNEL_METRICS:
            if mrow is None or metric_key in B2G_PURE_FACT_SKIP:
                continue
            v = monthly.cell(mrow, col).value
            val = coerce(v, pct)
            if val is None:
                continue
            rows.append(("b2g", "1", None, metric_key, val, "month", period_date))

    return rows


# ─────────────────────────────────────────────────────────────────────
# Weekly rollup: from daily rows, aggregate by ISO week (monday).
# Sum for cumulative metrics, average for non-cumulative.
# ─────────────────────────────────────────────────────────────────────

NON_CUMULATIVE_KEYS = {
    "total_ql2p_p", "total_ql2p_f", "total_avgCheck_p", "total_avgCheck_f",
    "buh_ql2p_p", "buh_ql2p_f", "buh_avgCheck_p", "buh_avgCheck_f",
    "med_ql2p_p", "med_ql2p_f", "med_avgCheck_p", "med_avgCheck_f",
    "calls_managersOnLine_p", "calls_managersOnLine_f",
    "calls_avgWait_p", "calls_avgWait_f",
    "calls_dialPercent_p", "calls_dialPercent_f",
    "calls_sla_p", "calls_sla_f",
    "okk_buh1_p", "okk_buh1_f", "okk_buh2_p", "okk_buh2_f",
    "okk_med1_p", "okk_med1_f", "okk_avg_p", "okk_avg_f",
    # B2G non-cumulative
    "qualLeadsPercent", "convQualTask", "convTaskConsult", "convConsultTerm",
    "dialPercent", "regulationPercent", "avgCallsPerLead", "sla_p", "sla_f",
    "sla_shift_f", "tlt_f", "okk_p", "okk_f", "roleplay_p", "roleplay_f",
    "staffCount", "managersOnLine", "avgPortfolio",
    "avgDialogPerEmployee", "avgDialogMinutes", "avgWait_p", "avgWait_f",
}


def compute_weekly(daily_rows: list[tuple]) -> list[tuple]:
    """Input: (dept, line, user_id, key, val, 'day', 'YYYY-MM-DD')
    Output: (dept, line, user_id, key, val, 'week', 'YYYY-Www')"""
    groups: dict[tuple, list[float]] = {}
    for dept, line, uid, key, val, ptype, pdate in daily_rows:
        if ptype != "day":
            continue
        try:
            num = float(val)
        except ValueError:
            continue
        d = date.fromisoformat(pdate)
        iso_year, iso_week, _ = d.isocalendar()
        period_date = f"{iso_year}-W{iso_week:02d}"
        groups.setdefault((dept, line, uid, key, period_date), []).append(num)

    out: list[tuple] = []
    for (dept, line, uid, key, pdate), vals in groups.items():
        if key in NON_CUMULATIVE_KEYS:
            num = sum(vals) / len(vals)
        else:
            num = sum(vals)
        if abs(num - round(num)) < 1e-9:
            s = str(int(round(num)))
        else:
            s = f"{num:.2f}".rstrip("0").rstrip(".")
        out.append((dept, line, uid, key, s, "week", pdate))
    return out


# ─────────────────────────────────────────────────────────────────────
# DB upsert
# ─────────────────────────────────────────────────────────────────────

PARTIAL_IDX_SQL = """
CREATE UNIQUE INDEX IF NOT EXISTS daily_plans_unique_null_user
ON daily_plans (department, line, metric_key, period_type, period_date)
WHERE user_id IS NULL
"""

UPSERT_SQL = """
INSERT INTO daily_plans (department, line, user_id, metric_key, plan_value,
                         period_type, period_date, created_at, updated_at)
VALUES %s
ON CONFLICT (department, line, metric_key, period_type, period_date)
WHERE user_id IS NULL
DO UPDATE SET plan_value = EXCLUDED.plan_value, updated_at = NOW()
"""


def upsert_rows(conn, rows: list[tuple]) -> None:
    cur = conn.cursor()
    cur.execute(PARTIAL_IDX_SQL)
    # All rows have user_id=None; strip to the 7 columns the SQL expects
    values = [
        (dept, line, uid, key, val, ptype, pdate, "NOW()", "NOW()")  # timestamps filled by DB
        for dept, line, uid, key, val, ptype, pdate in rows
    ]
    # Use a tuple of placeholders; created_at/updated_at use DEFAULT via DB
    # (we'll drop them from the INSERT target and let column defaults handle it)
    insert_sql = """
        INSERT INTO daily_plans (department, line, user_id, metric_key, plan_value,
                                 period_type, period_date)
        VALUES %s
        ON CONFLICT (department, line, metric_key, period_type, period_date)
        WHERE user_id IS NULL
        DO UPDATE SET plan_value = EXCLUDED.plan_value, updated_at = NOW()
    """
    stripped = [(d, l, u, k, v, t, pd) for (d, l, u, k, v, t, pd) in rows]
    execute_values(cur, insert_sql, stripped, page_size=1000)
    conn.commit()
    print(f"  UPSERTED: {cur.rowcount} rows touched")


def main() -> None:
    dry = "--dry" in sys.argv or "--dry-run" in sys.argv
    url = load_env()

    print(f"Extracting B2B from {B2B_FILE.name}…")
    b2b = extract_b2b_rows()
    print(f"  → {len(b2b)} rows (day+month)")

    print(f"Extracting B2G from {B2G_FILE.name}…")
    b2g = extract_b2g_rows()
    print(f"  → {len(b2g)} rows (day+month)")

    print("Computing weekly rollups…")
    weekly = compute_weekly(b2b) + compute_weekly(b2g)
    print(f"  → {len(weekly)} week rows")

    total = b2b + b2g + weekly
    print(f"\nTotal rows to upsert: {len(total)}")

    # Show summary per period_type / department
    summary: dict[tuple[str, str], int] = {}
    for r in total:
        summary[(r[0], r[5])] = summary.get((r[0], r[5]), 0) + 1
    for (dept, ptype), n in sorted(summary.items()):
        print(f"  {dept:4s} {ptype:6s}: {n}")

    if dry:
        print("\n[DRY RUN] no DB writes.")
        return

    print("\nConnecting to DB…")
    conn = psycopg2.connect(url)
    try:
        upsert_rows(conn, total)
    finally:
        conn.close()
    print("Done.")


if __name__ == "__main__":
    main()
